import json
import logging
import re
from dataclasses import dataclass
from typing import Any, List
from urllib.parse import urljoin

import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.hyperlink import Hyperlink

from config import BASE_URL
from logger import log_time


class HTTPResponseError(Exception):
    def __init__(self, url: str, status_code: int) -> None:
        super().__init__(
            f"Error getting page: {url}, status code: {status_code}"
        )


@dataclass(frozen=True)
class CourseLink:
    name: str
    link: str
    description: str


@dataclass(frozen=True)
class CourseModule:
    title: str
    description: str
    topics: List[str]


@dataclass(frozen=True)
class CourseDetail:
    modules: List[CourseModule]


@log_time
def fetch_full_page(url: str, timeout: int = 10) -> str:
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=chrome_options)
    driver.get(url)

    while click_show_more_button(driver, timeout):
        WebDriverWait(driver, timeout).until_not(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, ".CourseModulesBlock_showMoreButton__N0f0_")
            )
        )

    page_content = driver.page_source
    driver.quit()
    return page_content


def click_show_more_button(driver: webdriver.Chrome, timeout: int) -> bool:
    try:
        show_more_button = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, ".CourseModulesBlock_showMoreButton__N0f0_")
            )
        )
        driver.execute_script("arguments[0].click();", show_more_button)
        return True
    except Exception as e:
        logging.debug(f"No more 'Show more' button found: {e}")
        return False


def get_course_detail(url: str) -> (CourseDetail, int, int, str, str):
    content = fetch_full_page(url)
    soup = BeautifulSoup(content, "html.parser")

    num_modules, num_topics = extract_modules_and_topics(soup)
    full_time_duration, flex_time_duration = extract_durations(soup)
    modules = extract_modules(soup)

    return (
        CourseDetail(modules=modules),
        num_modules,
        num_topics,
        full_time_duration,
        flex_time_duration,
    )


def extract_modules_and_topics(soup: BeautifulSoup) -> (int, int):
    modules_heading = soup.find(
        "div",
        class_="CourseModulesHeading_headingGrid__ynoxV"
    )

    if modules_heading is None:
        logging.error("Modules heading not found.")
        raise ValueError("Modules heading not found.")

    num_modules = int(
        modules_heading.find(
            "div",
            class_="CourseModulesHeading_modulesNumber__UrnUh"
        )
        .text.strip()
        .split()[0]
    )
    num_topics = int(
        modules_heading.find(
            "div",
            class_="CourseModulesHeading_topicsNumber__5IA8Z"
        )
        .text.strip()
        .split()[0]
    )

    return num_modules, num_topics


def extract_durations(soup: BeautifulSoup) -> (str, str):
    full_time_duration = ""
    flex_time_duration = ""

    comparison_tables = soup.find_all(
        "div",
        class_=re.compile(r"ComparisonTable_wrapper__.*")
    )

    if not comparison_tables:
        logging.error("No comparison tables found.")
        raise ValueError("No comparison tables found.")

    for table in comparison_tables:
        cells = table.find_all("div", class_=re.compile(r"ComparisonTable_cell__.*"))

        for cell in cells:
            if cell.find("img", alt="Calendar"):
                full_time_duration = extract_duration_from_table(table)
            elif cell.find("img", alt="One o'clock"):
                flex_time_duration = extract_duration_from_table(table)

    return full_time_duration, flex_time_duration


def extract_table_header(table: BeautifulSoup) -> str:
    header = table.find("div", class_="ComparisonTable_row__P2dAA")

    if header:
        return header.text.strip()

    return ""


def extract_duration_from_table(table: BeautifulSoup) -> str:
    rows = table.find_all("div", class_="ComparisonTable_row__P2dAA")
    for row in rows:
        title = row.find(
            "div",
            class_="ComparisonTable_cell__8DNfm ComparisonTable_rowTitle__hwc7p"
        )
        if title and title.text.strip() == "Тривалість":
            value = row.find_all("div", class_="ComparisonTable_cell__8DNfm")[1]
            if value:
                return value.text.strip()
    return ""


def extract_modules(soup: BeautifulSoup) -> List[CourseModule]:
    modules = []
    module_items = soup.find_all(
        "div",
        class_=re.compile(r"CourseModuleItem_grid__.*")
    )

    for item in module_items:
        title, description = extract_module_title_and_description(item)
        topics = extract_module_topics(item)

        if title and description:
            modules.append(
                CourseModule(
                    title=title,
                    description=description,
                    topics=topics
                )
            )

    return modules


def extract_module_title_and_description(item: BeautifulSoup) -> (str, str):
    title_container = item.find(
        "div",
        class_=re.compile(r"CourseModuleItem_titleContainer__.*")
    )
    title = title_container.find(
        "h4",
        class_=re.compile(r"typography_landingH5__.*")
    ) if title_container else None
    description = item.find(
        "p",
        class_=re.compile(r"CourseModuleItem_description__.*")
    )

    if title and description:
        return title.text.strip(), description.text.strip()

    return "", ""


def extract_module_topics(item: BeautifulSoup) -> List[str]:
    topics_section = item.find(
        "div",
        class_="CourseModuleItem_topicsList__Dmm8g"
    )
    topics = []
    if topics_section:
        topics = [
            topic.text.strip()
            for topic in topics_section.find_all(
                "p", class_="typography_landingTextMain__Rc8BD"
            )
        ]

    return topics


def get_all_courses(url: str) -> list[CourseLink]:
    content = fetch_full_page(url)
    soup = BeautifulSoup(content, "html.parser")
    courses = []

    profession_cards = soup.find_all(
        "div", class_=re.compile(r"ProfessionCard_cardWrapper__JQBNJ")
    )

    for card in profession_cards:
        course = extract_course_from_card(card)
        if course:
            courses.append(course)

    return courses


def extract_course_from_card(card: BeautifulSoup) -> CourseLink | None:
    name_tag = card.find("a", class_=re.compile(r"ProfessionCard_title__.*"))
    description_tag = card.find(
        "p",
        class_=re.compile(r"ProfessionCard_subtitle__.*")
    )

    if name_tag and description_tag:
        name = name_tag.find("h3").text.strip()
        link = urljoin(BASE_URL, name_tag.get("href"))
        description = card.find(
            "p",
            class_=re.compile(r"typography_landingTextMain__.* mb-32")
        ).text.strip()
        return CourseLink(name=name, link=link, description=description)

    logging.warning(f"Missing course name or description in card: {card}")
    return None


def write_to_json(courses_details: list[dict], file_name: str) -> None:
    with open(file_name, "w", encoding="utf-8") as f:
        json.dump(courses_details, f, ensure_ascii=False, indent=4)
    logging.info(f"Write to file: {file_name}")


def sanitize_sheet_name(sheet_name: str) -> str:
    return re.sub(r'[\/:*?"<>|]', "_", sheet_name)[:31]


def add_line_breaks(text: str, max_words: int) -> str:
    words = text.split()
    lines = [
        " ".join(words[i: i + max_words])
        for i in range(0, len(words), max_words)
    ]
    return "\n".join(lines)


def write_to_excel(courses_data: list[dict], file_name: str) -> None:
    summary_data = create_summary_data(courses_data)
    df_summary = pd.DataFrame(summary_data)

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        for course in courses_data:
            write_course_details_to_sheet(writer, course)

    format_excel_file(file_name)
    logging.info(f"Write to Excel file: {file_name}")


def create_summary_data(courses_data: list[dict]) -> List[dict]:
    return [
        {
            "Name": course["name"],
            "Link": course["link"],
            "Description": course["description"],
            "Number of Modules": course["num_modules"],
            "Number of Topics": course["num_topics"],
            "Full-Time Duration": course["full_time_duration"],
            "Flex-Time Duration": course["flex_time_duration"],
        }
        for course in courses_data
    ]


def write_course_details_to_sheet(writer: pd.ExcelWriter, course: dict) -> None:
    sheet_name = sanitize_sheet_name(course["name"])
    module_data = create_module_data(course["details"]["modules"])
    df_modules = pd.DataFrame(module_data)
    df_modules.to_excel(writer, sheet_name=sheet_name, index=False)


def create_module_data(modules: List[dict]) -> List[dict]:
    return [
        {
            "Title": module["title"],
            "Description": module["description"],
            "Topics": ", ".join(module["topics"]),
        }
        for module in modules
    ]


def format_excel_file(file_name: str) -> None:
    workbook = load_workbook(file_name)
    format_summary_sheet(workbook)
    format_all_sheets(workbook)
    workbook.save(file_name)


def format_summary_sheet(workbook: Any) -> None:
    summary_sheet = workbook["Summary"]
    for row in summary_sheet.iter_rows(
            min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2
    ):
        for cell in row:
            link = cell.value
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=link)
            cell.style = "Hyperlink"
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )


def format_all_sheets(workbook: Any) -> None:
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        format_worksheet(worksheet)


def format_worksheet(worksheet: Any) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )

    for col in worksheet.columns:
        adjust_column_width(worksheet, col)


def adjust_column_width(worksheet: Any, col: Any) -> None:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except Exception:
            pass
    adjusted_width = min((max_length + 2), 50)
    worksheet.column_dimensions[column].width = adjusted_width
