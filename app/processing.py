import json
import logging
import re
from typing import Any, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.hyperlink import Hyperlink


def write_to_json(courses_details: list[dict], file_name: str) -> None:
    with open(file_name, "w", encoding="utf-8") as f:
        json.dump(courses_details, f, ensure_ascii=False, indent=4)
    logging.info(f"Write to file: {file_name}")


def sanitize_sheet_name(sheet_name: str) -> str:
    return re.sub(r'[\/:*?"<>|]', "_", sheet_name)[:31]


def add_line_breaks(text: str, max_words: int) -> str:
    words = text.split()
    lines = [
        " ".join(words[i: i + max_words])
        for i in range(0, len(words), max_words)
    ]
    return "\n".join(lines)


def write_to_excel(courses_data: list[dict], file_name: str) -> None:
    summary_data = create_summary_data(courses_data)
    df_summary = pd.DataFrame(summary_data)

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        for course in courses_data:
            write_course_details_to_sheet(writer, course)

    format_excel_file(file_name)
    logging.info(f"Write to Excel file: {file_name}")


def create_summary_data(courses_data: list[dict]) -> List[dict]:
    return [
        {
            "Name": course["name"],
            "Link": course["link"],
            "Description": course["description"],
            "Number of Modules": course["num_modules"],
            "Number of Topics": course["num_topics"],
            "Full-Time Duration": course["full_time_duration"],
            "Flex-Time Duration": course["flex_time_duration"],
        }
        for course in courses_data
    ]


def write_course_details_to_sheet(writer: pd.ExcelWriter, course: dict) -> None:
    sheet_name = sanitize_sheet_name(course["name"])
    module_data = create_module_data(course["details"]["modules"])
    df_modules = pd.DataFrame(module_data)
    df_modules.to_excel(writer, sheet_name=sheet_name, index=False)


def create_module_data(modules: List[dict]) -> List[dict]:
    return [
        {
            "Title": module["title"],
            "Description": module["description"],
            "Topics": ", ".join(module["topics"]),
        }
        for module in modules
    ]


def format_excel_file(file_name: str) -> None:
    workbook = load_workbook(file_name)
    format_summary_sheet(workbook)
    format_all_sheets(workbook)
    workbook.save(file_name)


def format_summary_sheet(workbook: Any) -> None:
    summary_sheet = workbook["Summary"]
    for row in summary_sheet.iter_rows(
            min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2
    ):
        for cell in row:
            link = cell.value
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=link)
            cell.style = "Hyperlink"
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )


def format_all_sheets(workbook: Any) -> None:
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        format_worksheet(worksheet)


def format_worksheet(worksheet: Any) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )

    for col in worksheet.columns:
        adjust_column_width(worksheet, col)


def adjust_column_width(worksheet: Any, col: Any) -> None:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except Exception:
            pass
    adjusted_width = min((max_length + 2), 50)
    worksheet.column_dimensions[column].width = adjusted_width
